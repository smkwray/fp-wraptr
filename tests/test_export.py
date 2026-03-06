"""Tests for workbook export utilities."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from fp_wraptr.analysis.export import export_comparison_workbook, export_run_workbook
from fp_wraptr.io.parser import ForecastVariable, FPOutputData
from fp_wraptr.scenarios.config import ScenarioConfig
from fp_wraptr.scenarios.runner import ScenarioResult


def _build_output() -> FPOutputData:
    return FPOutputData(
        forecast_start="2025.4",
        forecast_end="2025.6",
        periods=["2025.4", "2025.5", "2025.6"],
        variables={
            "GDPR": ForecastVariable(
                var_id=1,
                name="GDPR",
                levels=[100.0, 101.0, 102.0],
                changes=[1.0, 1.0, 1.0],
                pct_changes=[1.0, 1.0, 1.0],
            ),
            "UR": ForecastVariable(
                var_id=2,
                name="UR",
                levels=[4.0, 3.8, 3.6],
                changes=[-0.2, -0.2, -0.2],
                pct_changes=[-5.0, -5.2, -5.4],
            ),
        },
    )


def _scenario_result(tmp_path: Path, name: str, with_output: bool = True) -> ScenarioResult:
    return ScenarioResult(
        config=ScenarioConfig(name=name),
        output_dir=tmp_path / name,
        parsed_output=_build_output() if with_output else None,
    )


def test_export_run_workbook(tmp_path):
    result = _scenario_result(tmp_path, "run")
    output_path = tmp_path / "run.xlsx"

    exported = export_run_workbook(result, output_path)
    assert exported == output_path
    workbook = openpyxl.load_workbook(exported)
    assert set(workbook.sheetnames) == {"Config", "Forecast", "Changes", "PctChanges"}


def test_export_run_workbook_no_output(tmp_path):
    result = _scenario_result(tmp_path, "run", with_output=False)
    output_path = tmp_path / "run.xlsx"

    exported = export_run_workbook(result, output_path)
    workbook = openpyxl.load_workbook(exported)
    assert workbook.sheetnames == ["Config"]


def test_export_comparison_workbook(tmp_path):
    baseline = _scenario_result(tmp_path, "baseline")
    scenario = _scenario_result(tmp_path, "scenario")
    output_path = tmp_path / "comparison.xlsx"

    # create a difference so comparison sheet exists
    scenario.parsed_output.variables["GDPR"].levels[-1] = 103.0  # type: ignore[index]

    exported = export_comparison_workbook(baseline, scenario, output_path)
    workbook = openpyxl.load_workbook(exported)
    assert set(["Config", "Baseline", "Scenario", "Comparison"]).issubset(set(workbook.sheetnames))


def test_export_comparison_workbook_with_diff(tmp_path):
    baseline = _scenario_result(tmp_path, "baseline")
    scenario = _scenario_result(tmp_path, "scenario")
    output_path = tmp_path / "comparison.xlsx"
    diff_result = {
        "deltas": {
            "GDPR": {
                "baseline": 102.0,
                "scenario": 103.0,
                "abs_delta": 1.0,
                "pct_delta": 0.98,
            }
        }
    }

    exported = export_comparison_workbook(
        baseline,
        scenario,
        output_path,
        diff_result=diff_result,
    )
    workbook = openpyxl.load_workbook(exported)
    sheet = workbook["Comparison"]
    assert sheet.cell(row=2, column=1).value == "GDPR"
    assert sheet.cell(row=2, column=2).value == 102.0
